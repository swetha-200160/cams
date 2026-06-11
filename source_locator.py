from __future__ import annotations

import csv
from pathlib import Path
from typing import Any, Dict, Iterable, Optional

from models import EvidenceReference, FileLocator

try:
    from pypdf import PdfReader
except ImportError:
    PdfReader = None  # type: ignore[assignment,misc]

try:
    from docx import Document
except ImportError:
    Document = None  # type: ignore[assignment,misc]

try:
    from openpyxl import load_workbook
except ImportError:
    load_workbook = None  # type: ignore[assignment,misc]


class SourceLocatorService:
    def __init__(self, application_id: str, input_docs_dir: Path):
        self.application_id = application_id
        self.input_docs_dir = input_docs_dir

    def locate(
        self,
        document_name: Optional[str],
        source_field: Optional[str],
        source_year: Optional[str],
        extracted_value: Any,
        citation_id: str,
    ) -> EvidenceReference:
        document_name = document_name or "Unknown source"
        file_path = self._find_document(document_name)
        locator = FileLocator(type="unknown")
        excerpt: Optional[str] = None
        hyperlink = f"/api/files/{self.application_id}/{document_name}"

        if file_path and file_path.suffix.lower() == ".pdf":
            locator, excerpt = self._locate_pdf(file_path, source_field, source_year, extracted_value)
        elif file_path and file_path.suffix.lower() in {".xlsx", ".xlsm", ".xltx", ".csv"}:
            locator, excerpt = self._locate_sheet(file_path, source_field, source_year, extracted_value)
        elif file_path and file_path.suffix.lower() == ".docx":
            locator, excerpt = self._locate_docx(file_path, source_field, source_year, extracted_value)
        elif file_path and file_path.suffix.lower() in {".png", ".jpg", ".jpeg", ".webp"}:
            locator = FileLocator(type="file", label="Image evidence")
        elif file_path:
            locator = FileLocator(type="file", label="Download source file")

        return EvidenceReference(
            id=citation_id,
            document_name=document_name,
            document_path=str(file_path) if file_path else None,
            hyperlink=hyperlink,
            excerpt=excerpt,
            source_field=source_field,
            source_year=source_year,
            extracted_value=extracted_value,
            locator=locator,
        )

    def _find_document(self, document_name: str) -> Optional[Path]:
        if not self.input_docs_dir.exists():
            return None
        exact = self.input_docs_dir / document_name
        if exact.exists():
            return exact
        lowered = document_name.lower()
        for path in self.input_docs_dir.rglob("*"):
            if path.is_file() and path.name.lower() == lowered:
                return path
        for path in self.input_docs_dir.rglob("*"):
            if path.is_file() and lowered in path.name.lower():
                return path
        return None

    def _locate_pdf(self, file_path: Path, source_field: Optional[str], source_year: Optional[str], extracted_value: Any):
        if PdfReader is None:
            return FileLocator(type="file", label="PDF available"), None

        try:
            header = file_path.read_bytes()[:5]
            if not header.startswith(b"%PDF"):
                return FileLocator(type="file", label="PDF source"), None
            reader = PdfReader(str(file_path))
            needles = [self._stringify(extracted_value), source_field, source_year]
            needles = [n.lower() for n in needles if n]
            for index, page in enumerate(reader.pages, start=1):
                text = (page.extract_text() or "").replace("\x00", " ")
                lowered = text.lower()
                if not needles or any(n in lowered for n in needles):
                    excerpt = self._excerpt(text, needles)
                    return FileLocator(type="page", page=index, label=f"Page {index}"), excerpt
        except Exception:
            return FileLocator(type="file", label="PDF source"), None
        return FileLocator(type="file", label="PDF source"), None

    def _locate_docx(self, file_path: Path, source_field: Optional[str], source_year: Optional[str], extracted_value: Any):
        if Document is None:
            return FileLocator(type="file", label="Word document available"), None

        try:
            doc = Document(str(file_path))
            needles = [self._stringify(extracted_value), source_field, source_year]
            needles = [n.lower() for n in needles if n]
            paragraphs = [p.text for p in doc.paragraphs]
            for idx, para in enumerate(paragraphs):
                lowered = para.lower()
                if not needles or any(n in lowered for n in needles):
                    return FileLocator(type="paragraph", paragraph_index=idx, label=f"Paragraph {idx + 1}"), para[:500]
        except Exception:
            return FileLocator(type="file", label="Word source"), None
        return FileLocator(type="file", label="Word source"), None

    def _locate_sheet(self, file_path: Path, source_field: Optional[str], source_year: Optional[str], extracted_value: Any):
        if file_path.suffix.lower() == ".csv":
            return self._locate_csv(file_path, source_field, source_year, extracted_value)
        return self._locate_xlsx(file_path, source_field, source_year, extracted_value)

    def _locate_xlsx(self, file_path: Path, source_field: Optional[str], source_year: Optional[str], extracted_value: Any):
        if load_workbook is None:
            return FileLocator(type="file", label="Spreadsheet available"), None

        try:
            wb = load_workbook(file_path, read_only=True, data_only=True)
            target = self._stringify(extracted_value)
            year = self._stringify(source_year)
            for ws in wb.worksheets:
                rows = list(ws.iter_rows(values_only=True))
                for idx, row in enumerate(rows, start=1):
                    row_values = {self._excel_col(i + 1): value for i, value in enumerate(row)}
                    if self._row_matches(row_values.values(), target, source_field, year):
                        cell = self._find_matching_cell(row_values, target, source_field, year)
                        before = {self._excel_col(i + 1): value for i, value in enumerate(rows[idx - 2])} if idx > 1 else None
                        after = {self._excel_col(i + 1): value for i, value in enumerate(rows[idx])} if idx < len(rows) else None
                        return (
                            FileLocator(
                                type="sheet_cell" if cell else "sheet_row",
                                sheet_name=ws.title,
                                cell=cell,
                                row_number=idx,
                                before_row=before,
                                current_row=row_values,
                                after_row=after,
                                label=f"{ws.title} row {idx}" + (f", cell {cell}" if cell else ""),
                            ),
                            self._sheet_excerpt(ws.title, idx, row_values, source_field),
                        )
        except Exception:
            return FileLocator(type="file", label="Spreadsheet source"), None
        return FileLocator(type="file", label="Spreadsheet source"), None

    def _locate_csv(self, file_path: Path, source_field: Optional[str], source_year: Optional[str], extracted_value: Any):
        target = self._stringify(extracted_value)
        year = self._stringify(source_year)
        try:
            with open(file_path, encoding="utf-8", errors="ignore") as fh:
                reader = list(csv.reader(fh))
            for idx, row in enumerate(reader, start=1):
                if self._row_matches(row, target, source_field, year):
                    row_map = {self._excel_col(i + 1): value for i, value in enumerate(row)}
                    before = {self._excel_col(i + 1): value for i, value in enumerate(reader[idx - 2])} if idx > 1 else None
                    after = {self._excel_col(i + 1): value for i, value in enumerate(reader[idx])} if idx < len(reader) else None
                    cell = self._find_matching_cell(row_map, target, source_field, year)
                    return (
                        FileLocator(
                            type="sheet_cell" if cell else "sheet_row",
                            sheet_name=file_path.name,
                            cell=cell,
                            row_number=idx,
                            before_row=before,
                            current_row=row_map,
                            after_row=after,
                            label=f"{file_path.name} row {idx}" + (f", cell {cell}" if cell else ""),
                        ),
                        self._sheet_excerpt(file_path.name, idx, row_map, source_field),
                    )
        except Exception:
            return FileLocator(type="file", label="Spreadsheet source"), None
        return FileLocator(type="file", label="Spreadsheet source"), None

    @staticmethod
    def _row_matches(values: Iterable[Any], target: Optional[str], field_name: Optional[str], year: Optional[str]) -> bool:
        normalized = [SourceLocatorService._stringify(v).lower() for v in values if v not in (None, "")]
        if target and any(target.lower() == value or target.lower() in value for value in normalized):
            return True
        if field_name and any(field_name.lower() in value for value in normalized):
            return True
        if year and any(year.lower() == value or year.lower() in value for value in normalized):
            return True
        return False

    @staticmethod
    def _find_matching_cell(row_map: Dict[str, Any], target: Optional[str], field_name: Optional[str], year: Optional[str]) -> Optional[str]:
        pairs = [(target, True), (field_name, False), (year, False)]
        for needle, exact_first in pairs:
            if not needle:
                continue
            lowered = needle.lower()
            for col, value in row_map.items():
                value_text = SourceLocatorService._stringify(value).lower()
                if exact_first and value_text == lowered:
                    return col
                if lowered and lowered in value_text:
                    return col
        return None

    @staticmethod
    def _sheet_excerpt(sheet_name: str, row_idx: int, row_map: Dict[str, Any], source_field: Optional[str]) -> str:
        preview = ", ".join(f"{col}={value}" for col, value in list(row_map.items())[:8])
        field_label = f" for {source_field}" if source_field else ""
        return f"{sheet_name} row {row_idx}{field_label}: {preview}"

    @staticmethod
    def _excerpt(text: str, needles: list[str]) -> str:
        lowered = text.lower()
        for needle in needles:
            idx = lowered.find(needle)
            if idx >= 0:
                start = max(0, idx - 120)
                end = min(len(text), idx + 220)
                return text[start:end].replace("\n", " ")
        return text[:340].replace("\n", " ")

    @staticmethod
    def _stringify(value: Any) -> Optional[str]:
        if value is None:
            return None
        if isinstance(value, float):
            if value.is_integer():
                return str(int(value))
            # Use up to 10 significant digits so values like 1.85714 are
            # preserved accurately when searching inside source documents.
            return f"{value:.10g}"
        return str(value).strip()

    @staticmethod
    def _excel_col(index: int) -> str:
        result = ""
        while index:
            index, remainder = divmod(index - 1, 26)
            result = chr(65 + remainder) + result
        return result
